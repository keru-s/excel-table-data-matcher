from __future__ import annotations

import tempfile
import unittest
from unittest import mock
from pathlib import Path

import openpyxl

from excel_core import (
    DEDUP_STRATEGY_BOTH_KEYS,
    DEDUP_STRATEGY_NONE,
    ExportConfig,
    LEFT_SOURCE,
    RIGHT_SOURCE,
    MatchPair,
    OutputColumn,
    export_matches,
    load_preview,
)


class ExcelCoreTest(unittest.TestCase):
    def test_load_preview_skips_empty_first_sheet(self) -> None:
        with tempfile.TemporaryDirectory(prefix="excel_core_sheet_") as temp_dir:
            workbook_path = Path(temp_dir) / "sheet_switch.xlsx"
            workbook = openpyxl.Workbook()
            empty_sheet = workbook.active
            empty_sheet.title = "空表"
            data_sheet = workbook.create_sheet("有效数据")
            data_sheet.append(["id", "name"])
            data_sheet.append([1, "Alice"])
            workbook.save(workbook_path)
            workbook.close()

            preview = load_preview(str(workbook_path), None, "utf-8")

            self.assertEqual(preview.selected_sheet, "有效数据")
            self.assertEqual([column.raw_name for column in preview.columns], ["id", "name"])
            self.assertEqual(preview.rows, [["1", "Alice"]])

    def test_export_matches_keeps_first_row_per_match_key(self) -> None:
        with tempfile.TemporaryDirectory(prefix="excel_core_dedupe_") as temp_dir:
            temp_path = Path(temp_dir)
            left_path = temp_path / "left.xlsx"
            right_path = temp_path / "right.xlsx"

            left_wb = openpyxl.Workbook()
            left_ws = left_wb.active
            left_ws.title = "源表"
            left_ws.append(["id", "name"])
            left_ws.append([1, "Alice"])
            left_ws.append([1, "Bob"])
            left_wb.save(left_path)
            left_wb.close()

            right_wb = openpyxl.Workbook()
            right_ws = right_wb.active
            right_ws.title = "目标表"
            right_ws.append(["user_id", "city"])
            right_ws.append([1, "Shanghai"])
            right_ws.append([1, "Shanghai"])
            right_ws.append([1, "Beijing"])
            right_wb.save(right_path)
            right_wb.close()

            output_path = export_matches(
                ExportConfig(
                    left_path=str(left_path),
                    right_path=str(right_path),
                    left_sheet="源表",
                    right_sheet="目标表",
                    encoding="utf-8",
                    match_pairs=[MatchPair("id", "user_id", "A · id", "A · user_id")],
                    output_columns=[
                        OutputColumn(LEFT_SOURCE, "name", "B · name"),
                        OutputColumn(RIGHT_SOURCE, "city", "C · city · 被匹配文件"),
                    ],
                    output_filename="dedupe_result",
                )
            )

            output_wb = openpyxl.load_workbook(output_path, read_only=True)
            try:
                rows = list(output_wb.active.iter_rows(values_only=True))
            finally:
                output_wb.close()

            self.assertEqual(
                rows,
                [
                    ("A · id", "B · name", "C · city · 被匹配文件"),
                    (1, "Alice", "Shanghai"),
                    (1, "Alice", "Beijing"),
                ],
            )

    def test_export_matches_deduplicates_by_multiple_match_columns(self) -> None:
        with tempfile.TemporaryDirectory(prefix="excel_core_multi_key_") as temp_dir:
            temp_path = Path(temp_dir)
            left_path = temp_path / "left.xlsx"
            right_path = temp_path / "right.xlsx"

            left_wb = openpyxl.Workbook()
            left_ws = left_wb.active
            left_ws.title = "源表"
            left_ws.append(["city", "district", "name"])
            left_ws.append(["Guangzhou", "Tianhe", "Alice"])
            left_ws.append(["Guangzhou", "Tianhe", "Bob"])
            left_ws.append(["Guangzhou", "Yuexiu", "Carol"])
            left_wb.save(left_path)
            left_wb.close()

            right_wb = openpyxl.Workbook()
            right_ws = right_wb.active
            right_ws.title = "目标表"
            right_ws.append(["city", "district", "code"])
            right_ws.append(["Guangzhou", "Tianhe", "TH-001"])
            right_ws.append(["Guangzhou", "Tianhe", "TH-002"])
            right_ws.append(["Guangzhou", "Yuexiu", "YX-001"])
            right_wb.save(right_path)
            right_wb.close()

            output_path = export_matches(
                ExportConfig(
                    left_path=str(left_path),
                    right_path=str(right_path),
                    left_sheet="源表",
                    right_sheet="目标表",
                    encoding="utf-8",
                    match_pairs=[
                        MatchPair("city", "city", "A · city", "A · city"),
                        MatchPair("district", "district", "B · district", "B · district"),
                    ],
                    output_columns=[
                        OutputColumn(LEFT_SOURCE, "name", "C · name"),
                        OutputColumn(RIGHT_SOURCE, "code", "D · code · 被匹配文件"),
                    ],
                    output_filename="multi_key_result",
                )
            )

            output_wb = openpyxl.load_workbook(output_path, read_only=True)
            try:
                rows = list(output_wb.active.iter_rows(values_only=True))
            finally:
                output_wb.close()

            self.assertEqual(
                rows,
                [
                    ("A · city", "B · district", "C · name", "D · code · 被匹配文件"),
                    ("Guangzhou", "Tianhe", "Alice", "TH-001"),
                    ("Guangzhou", "Tianhe", "Alice", "TH-002"),
                    ("Guangzhou", "Yuexiu", "Carol", "YX-001"),
                ],
            )

    def test_export_matches_can_deduplicate_both_sides_by_match_key(self) -> None:
        with tempfile.TemporaryDirectory(prefix="excel_core_both_keys_") as temp_dir:
            temp_path = Path(temp_dir)
            left_path = temp_path / "left.xlsx"
            right_path = temp_path / "right.xlsx"

            left_wb = openpyxl.Workbook()
            left_ws = left_wb.active
            left_ws.title = "源表"
            left_ws.append(["city", "name"])
            left_ws.append(["Guangzhou", "Alice"])
            left_ws.append(["Guangzhou", "Bob"])
            left_wb.save(left_path)
            left_wb.close()

            right_wb = openpyxl.Workbook()
            right_ws = right_wb.active
            right_ws.title = "目标表"
            right_ws.append(["city", "code"])
            right_ws.append(["Guangzhou", "TH-001"])
            right_ws.append(["Guangzhou", "TH-002"])
            right_wb.save(right_path)
            right_wb.close()

            output_path = export_matches(
                ExportConfig(
                    left_path=str(left_path),
                    right_path=str(right_path),
                    left_sheet="源表",
                    right_sheet="目标表",
                    encoding="utf-8",
                    match_pairs=[MatchPair("city", "city", "A · city", "A · city")],
                    output_columns=[
                        OutputColumn(LEFT_SOURCE, "name", "B · name"),
                        OutputColumn(RIGHT_SOURCE, "code", "C · code · 被匹配文件"),
                    ],
                    output_filename="both_keys_result",
                    dedup_strategy=DEDUP_STRATEGY_BOTH_KEYS,
                )
            )

            output_wb = openpyxl.load_workbook(output_path, read_only=True)
            try:
                rows = list(output_wb.active.iter_rows(values_only=True))
            finally:
                output_wb.close()

            self.assertEqual(
                rows,
                [
                    ("A · city", "B · name", "C · code · 被匹配文件"),
                    ("Guangzhou", "Alice", "TH-001"),
                ],
            )

    def test_export_matches_can_disable_deduplication(self) -> None:
        with tempfile.TemporaryDirectory(prefix="excel_core_no_dedupe_") as temp_dir:
            temp_path = Path(temp_dir)
            left_path = temp_path / "left.xlsx"
            right_path = temp_path / "right.xlsx"

            left_wb = openpyxl.Workbook()
            left_ws = left_wb.active
            left_ws.title = "源表"
            left_ws.append(["city", "name"])
            left_ws.append(["Guangzhou", "Alice"])
            left_ws.append(["Guangzhou", "Bob"])
            left_wb.save(left_path)
            left_wb.close()

            right_wb = openpyxl.Workbook()
            right_ws = right_wb.active
            right_ws.title = "目标表"
            right_ws.append(["city", "code"])
            right_ws.append(["Guangzhou", "TH-001"])
            right_ws.append(["Guangzhou", "TH-002"])
            right_wb.save(right_path)
            right_wb.close()

            output_path = export_matches(
                ExportConfig(
                    left_path=str(left_path),
                    right_path=str(right_path),
                    left_sheet="源表",
                    right_sheet="目标表",
                    encoding="utf-8",
                    match_pairs=[MatchPair("city", "city", "A · city", "A · city")],
                    output_columns=[
                        OutputColumn(LEFT_SOURCE, "name", "B · name"),
                        OutputColumn(RIGHT_SOURCE, "code", "C · code · 被匹配文件"),
                    ],
                    output_filename="no_dedupe_result",
                    dedup_strategy=DEDUP_STRATEGY_NONE,
                )
            )

            output_wb = openpyxl.load_workbook(output_path, read_only=True)
            try:
                rows = list(output_wb.active.iter_rows(values_only=True))
            finally:
                output_wb.close()

            self.assertEqual(
                rows,
                [
                    ("A · city", "B · name", "C · code · 被匹配文件"),
                    ("Guangzhou", "Alice", "TH-001"),
                    ("Guangzhou", "Alice", "TH-002"),
                    ("Guangzhou", "Bob", "TH-001"),
                    ("Guangzhou", "Bob", "TH-002"),
                ],
            )

    def test_export_matches_fails_fast_when_estimated_rows_exceed_excel_limit(self) -> None:
        with tempfile.TemporaryDirectory(prefix="excel_core_limit_") as temp_dir:
            temp_path = Path(temp_dir)
            left_path = temp_path / "left.xlsx"
            right_path = temp_path / "right.xlsx"

            left_wb = openpyxl.Workbook()
            left_ws = left_wb.active
            left_ws.title = "源表"
            left_ws.append(["city", "name"])
            left_ws.append(["Guangzhou", "Alice"])
            left_ws.append(["Shenzhen", "Bob"])
            left_ws.append(["Beijing", "Carol"])
            left_ws.append(["Hangzhou", "David"])
            left_wb.save(left_path)
            left_wb.close()

            right_wb = openpyxl.Workbook()
            right_ws = right_wb.active
            right_ws.title = "目标表"
            right_ws.append(["city", "status"])
            right_ws.append(["Guangzhou", "启用"])
            right_ws.append(["Shenzhen", "启用"])
            right_ws.append(["Beijing", "启用"])
            right_ws.append(["Hangzhou", "启用"])
            right_wb.save(right_path)
            right_wb.close()

            with mock.patch("excel_core.EXCEL_MAX_ROWS_PER_SHEET", 3):
                with self.assertRaisesRegex(ValueError, "已超过 Excel 单个工作表的上限"):
                    export_matches(
                        ExportConfig(
                            left_path=str(left_path),
                            right_path=str(right_path),
                            left_sheet="源表",
                            right_sheet="目标表",
                            encoding="utf-8",
                            match_pairs=[MatchPair("city", "city", "A · city", "A · city")],
                            output_columns=[
                                OutputColumn(LEFT_SOURCE, "name", "B · name"),
                                OutputColumn(RIGHT_SOURCE, "status", "C · status · 被匹配文件"),
                            ],
                            output_filename="overflow_result",
                        )
                    )


if __name__ == "__main__":
    unittest.main()
