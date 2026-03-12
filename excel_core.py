from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import os
import sqlite3
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterator, Sequence

import chardet
import openpyxl
import xlrd
from openpyxl import Workbook


CSV_SHEET_NAME = "CSV文件"
LEFT_SOURCE = "left"
RIGHT_SOURCE = "right"
EXCEL_MAX_ROWS_PER_SHEET = 1_048_576
LOOKUP_IN_MEMORY_MAX_ROWS = 100_000
DEDUP_STRATEGY_LEFT_KEY_RIGHT_ROW = "left_key_right_row"
DEDUP_STRATEGY_NONE = "none"
DEDUP_STRATEGY_BOTH_KEYS = "both_keys"


@dataclass(frozen=True)
class ColumnOption:
    raw_name: str
    display_name: str


@dataclass(frozen=True)
class PreviewData:
    sheet_names: list[str]
    selected_sheet: str
    columns: list[ColumnOption]
    rows: list[list[str]]
    total_rows: int
    used_encoding: str


@dataclass(frozen=True)
class MatchPair:
    left_raw: str
    right_raw: str
    left_label: str
    right_label: str


@dataclass(frozen=True)
class OutputColumn:
    source: str
    raw_name: str
    label: str


@dataclass(frozen=True)
class ExportConfig:
    left_path: str
    right_path: str
    left_sheet: str
    right_sheet: str
    encoding: str
    match_pairs: Sequence[MatchPair]
    output_columns: Sequence[OutputColumn]
    output_filename: str
    dedup_strategy: str = DEDUP_STRATEGY_LEFT_KEY_RIGHT_ROW


@dataclass
class RowStream:
    headers: list[str]
    sheet_name: str
    total_rows: int
    used_encoding: str
    iterator: Iterator[dict[str, Any]]
    close: Callable[[], None]


class ExportCancelled(RuntimeError):
    pass


def load_preview(file_path: str, sheet_name: str | None, encoding: str, preview_rows: int = 12) -> PreviewData:
    sheet_names = list_sheet_names(file_path, encoding)
    stream = open_row_stream(file_path, sheet_name, encoding)
    rows: list[list[str]] = []
    try:
        for index, row in enumerate(stream.iterator):
            if index >= preview_rows:
                break
            rows.append([format_cell(row.get(header)) for header in stream.headers])
    finally:
        stream.close()

    columns = [
        ColumnOption(raw_name=header, display_name=f"{excel_column_name(idx)} · {header}")
        for idx, header in enumerate(stream.headers, start=1)
    ]
    return PreviewData(
        sheet_names=sheet_names,
        selected_sheet=stream.sheet_name,
        columns=columns,
        rows=rows,
        total_rows=stream.total_rows,
        used_encoding=stream.used_encoding,
    )


def list_sheet_names(file_path: str, encoding: str) -> list[str]:
    file_ext = Path(file_path).suffix.lower()
    if file_ext == ".csv":
        return [CSV_SHEET_NAME]
    if file_ext == ".xls":
        workbook = _open_xls_workbook(file_path)
        return workbook.sheet_names()
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def count_data_rows(file_path: str, sheet_name: str | None, encoding: str) -> int:
    file_ext = Path(file_path).suffix.lower()
    if file_ext == ".csv":
        used_encoding = choose_csv_encoding(file_path, encoding)
        with open(file_path, "r", encoding=used_encoding, newline="") as handle:
            reader = csv.reader(handle)
            next(reader, None)
            return sum(1 for _ in reader)
    if file_ext == ".xls":
        workbook = _open_xls_workbook(file_path)
        actual_sheet = _resolve_non_empty_xls_sheet(workbook, sheet_name)
        sheet = workbook.sheet_by_name(actual_sheet)
        return max(sheet.nrows - 1, 0)

    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        actual_sheet = _resolve_non_empty_xlsx_sheet(workbook, sheet_name)
        worksheet = workbook[actual_sheet]
        return max((worksheet.max_row or 0) - 1, 0)
    finally:
        workbook.close()


def export_matches(
    config: ExportConfig,
    progress_callback: Callable[[int, str], None] | None = None,
    is_cancelled: Callable[[], bool] | None = None,
) -> str:
    _emit_progress(progress_callback, 2, "正在分析文件...")

    left_rows = count_data_rows(config.left_path, config.left_sheet, config.encoding)
    right_rows = count_data_rows(config.right_path, config.right_sheet, config.encoding)
    output_path = build_output_path(config.right_path, config.output_filename)

    _emit_progress(
        progress_callback,
        8,
        f"匹配源文件 {left_rows:,} 行，被匹配文件 {right_rows:,} 行",
    )

    if not config.match_pairs:
        raise ValueError("至少需要设置一组匹配列。")

    lookup_columns = [column for column in config.output_columns if column.source == RIGHT_SOURCE]
    lookup = _create_lookup_backend(right_rows)
    right_unique_counts: dict[tuple[str, ...], int] = {}
    right_stream = open_row_stream(config.right_path, config.right_sheet, config.encoding)
    dedupe_left_by_key = config.dedup_strategy in (DEDUP_STRATEGY_LEFT_KEY_RIGHT_ROW, DEDUP_STRATEGY_BOTH_KEYS)
    dedupe_right_by_key = config.dedup_strategy == DEDUP_STRATEGY_BOTH_KEYS
    dedupe_right_by_row = config.dedup_strategy == DEDUP_STRATEGY_LEFT_KEY_RIGHT_ROW

    try:
        total_right = max(right_stream.total_rows, 1)
        for index, row in enumerate(right_stream.iterator, start=1):
            _ensure_not_cancelled(is_cancelled)
            key = tuple(normalize_match_value(row.get(pair.right_raw)) for pair in config.match_pairs)
            payload = {column.raw_name: row.get(column.raw_name) for column in lookup_columns}
            row_signature = _build_row_signature(row)
            if lookup.add(
                key,
                payload,
                row_signature=row_signature,
                dedupe_by_key=dedupe_right_by_key,
                dedupe_by_row=dedupe_right_by_row,
            ):
                right_unique_counts[key] = right_unique_counts.get(key, 0) + 1
            if index % 1000 == 0 or index == total_right:
                progress = 8 + int(index / total_right * 30)
                _emit_progress(progress_callback, progress, f"正在建立匹配索引 {index:,}/{right_stream.total_rows:,}")
    finally:
        right_stream.close()

    _emit_progress(progress_callback, 39, "正在评估结果规模...")
    estimated_rows = _estimate_output_rows(
        config,
        right_unique_counts,
        is_cancelled,
        dedupe_left_by_key=dedupe_left_by_key,
    )
    if estimated_rows > EXCEL_MAX_ROWS_PER_SHEET:
        lookup.close()
        raise ValueError(
            "当前匹配条件预计生成 "
            f"{estimated_rows:,} 行结果，已超过 Excel 单个工作表的上限 {EXCEL_MAX_ROWS_PER_SHEET:,} 行。"
            "请增加匹配列、缩小范围，或减少重复键后再导出。"
        )

    _emit_progress(progress_callback, 40, "正在导出匹配结果...")

    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="匹配结果")
    headers = [pair.left_label for pair in config.match_pairs] + [column.label for column in config.output_columns]
    worksheet.append(headers)

    written_rows = 0
    left_stream = open_row_stream(config.left_path, config.left_sheet, config.encoding)
    try:
        total_left = max(left_stream.total_rows, 1)
        seen_left_keys: set[tuple[str, ...]] = set()
        for index, row in enumerate(left_stream.iterator, start=1):
            _ensure_not_cancelled(is_cancelled)
            key = tuple(normalize_match_value(row.get(pair.left_raw)) for pair in config.match_pairs)
            if dedupe_left_by_key:
                if key in seen_left_keys:
                    continue
                seen_left_keys.add(key)
            matches = lookup.find(key)
            if matches:
                for matched_row in matches:
                    output_row = [prepare_output_value(row.get(pair.left_raw)) for pair in config.match_pairs]
                    for column in config.output_columns:
                        source_row = row if column.source == LEFT_SOURCE else matched_row
                        output_row.append(prepare_output_value(source_row.get(column.raw_name)))
                    worksheet.append(output_row)
                    written_rows += 1

            if index % 1000 == 0 or index == total_left:
                progress = 40 + int(index / total_left * 55)
                _emit_progress(
                    progress_callback,
                    progress,
                    f"已扫描匹配源文件 {index:,}/{left_stream.total_rows:,}，输出 {written_rows:,} 行",
                )
    finally:
        left_stream.close()
        lookup.close()

    _ensure_not_cancelled(is_cancelled)
    _emit_progress(progress_callback, 97, "正在保存结果文件...")
    workbook.save(output_path)
    workbook.close()
    _emit_progress(progress_callback, 100, f"导出完成，共输出 {written_rows:,} 行")
    return output_path


def build_output_path(reference_path: str, output_filename: str) -> str:
    target_dir = os.path.dirname(reference_path) or os.getcwd()
    safe_name = (output_filename or "匹配结果").strip() or "匹配结果"
    base_path = os.path.join(target_dir, f"{safe_name}.xlsx")
    if not os.path.exists(base_path):
        return os.path.normpath(base_path)

    timestamp = dt.datetime.now().strftime("%m%d_%H%M%S")
    return os.path.normpath(os.path.join(target_dir, f"{safe_name}_{timestamp}.xlsx"))


def open_row_stream(file_path: str, sheet_name: str | None, encoding: str) -> RowStream:
    file_ext = Path(file_path).suffix.lower()
    if file_ext == ".csv":
        return _open_csv_stream(file_path, encoding)
    if file_ext == ".xls":
        return _open_xls_stream(file_path, sheet_name)
    return _open_xlsx_stream(file_path, sheet_name)


def choose_csv_encoding(file_path: str, preferred: str) -> str:
    candidates = [preferred, "utf-8-sig", "utf-8", "gb18030", "gbk"]
    seen: set[str] = set()
    for candidate in candidates:
        if candidate in seen:
            continue
        seen.add(candidate)
        try:
            with open(file_path, "r", encoding=candidate, newline="") as handle:
                handle.readline()
            return candidate
        except UnicodeDecodeError:
            continue

    detected = chardet.detect(Path(file_path).read_bytes()[:65536]).get("encoding")
    if detected:
        try:
            with open(file_path, "r", encoding=detected, newline="") as handle:
                handle.readline()
            return detected
        except UnicodeDecodeError:
            pass
    return preferred


def excel_column_name(index: int) -> str:
    if index <= 0:
        return ""
    letters: list[str] = []
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def normalize_match_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.15g}"
    return str(value).strip()


def prepare_output_value(value: Any) -> Any:
    if isinstance(value, dt.datetime):
        if value.time() == dt.time():
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    return value


def format_cell(value: Any) -> str:
    output_value = prepare_output_value(value)
    if output_value is None:
        return ""
    return str(output_value)


def _open_csv_stream(file_path: str, encoding: str) -> RowStream:
    used_encoding = choose_csv_encoding(file_path, encoding)
    handle = open(file_path, "r", encoding=used_encoding, newline="")
    reader = csv.reader(handle)
    header_row = next(reader, [])
    headers = _make_unique_headers(header_row)

    def iterator() -> Iterator[dict[str, Any]]:
        for row in reader:
            yield dict(zip(headers, _normalize_row_values(row, len(headers))))

    total_rows = count_data_rows(file_path, CSV_SHEET_NAME, used_encoding)
    return RowStream(
        headers=headers,
        sheet_name=CSV_SHEET_NAME,
        total_rows=total_rows,
        used_encoding=used_encoding,
        iterator=iterator(),
        close=handle.close,
    )


def _open_xlsx_stream(file_path: str, sheet_name: str | None) -> RowStream:
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    actual_sheet = _resolve_non_empty_xlsx_sheet(workbook, sheet_name)
    worksheet = workbook[actual_sheet]
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, ())
    headers = _make_unique_headers(header_row)
    total_rows = max((worksheet.max_row or 0) - 1, 0)

    def iterator() -> Iterator[dict[str, Any]]:
        for row in rows:
            yield dict(zip(headers, _normalize_row_values(list(row), len(headers))))

    return RowStream(
        headers=headers,
        sheet_name=actual_sheet,
        total_rows=total_rows,
        used_encoding="utf-8",
        iterator=iterator(),
        close=workbook.close,
    )


def _open_xls_stream(file_path: str, sheet_name: str | None) -> RowStream:
    workbook = _open_xls_workbook(file_path)
    actual_sheet = _resolve_non_empty_xls_sheet(workbook, sheet_name)
    sheet = workbook.sheet_by_name(actual_sheet)
    headers = _make_unique_headers(sheet.row_values(0) if sheet.nrows else [])
    total_rows = max(sheet.nrows - 1, 0)

    def iterator() -> Iterator[dict[str, Any]]:
        for row_index in range(1, sheet.nrows):
            row = sheet.row_values(row_index)
            normalized = []
            for column_index, value in enumerate(_normalize_row_values(row, len(headers))):
                cell_type = sheet.cell_type(row_index, column_index) if column_index < sheet.ncols else xlrd.XL_CELL_EMPTY
                if cell_type == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                    except (TypeError, ValueError):
                        pass
                normalized.append(value)
            yield dict(zip(headers, normalized))

    return RowStream(
        headers=headers,
        sheet_name=actual_sheet,
        total_rows=total_rows,
        used_encoding="gb18030",
        iterator=iterator(),
        close=lambda: None,
    )


def _make_unique_headers(header_values: Sequence[Any]) -> list[str]:
    counts: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(header_values, start=1):
        base_name = str(value).strip() if value not in (None, "") else f"未命名列{index}"
        current_count = counts.get(base_name, 0) + 1
        counts[base_name] = current_count
        headers.append(base_name if current_count == 1 else f"{base_name} ({current_count})")
    return headers


def _normalize_row_values(values: Sequence[Any], width: int) -> list[Any]:
    padded = list(values[:width])
    if len(padded) < width:
        padded.extend([None] * (width - len(padded)))
    return padded


def _resolve_sheet_name(sheet_names: Sequence[str], requested: str | None) -> str:
    if requested and requested in sheet_names:
        return requested
    if not sheet_names:
        raise ValueError("文件中没有可读取的工作表。")
    return sheet_names[0]


def _resolve_non_empty_xlsx_sheet(workbook: openpyxl.Workbook, requested: str | None) -> str:
    if requested and requested in workbook.sheetnames:
        return requested
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if _xlsx_sheet_has_data(worksheet):
            return sheet_name
    return _resolve_sheet_name(workbook.sheetnames, requested)


def _resolve_non_empty_xls_sheet(workbook: xlrd.book.Book, requested: str | None) -> str:
    sheet_names = workbook.sheet_names()
    if requested and requested in sheet_names:
        return requested
    for sheet_name in sheet_names:
        sheet = workbook.sheet_by_name(sheet_name)
        if _xls_sheet_has_data(sheet):
            return sheet_name
    return _resolve_sheet_name(sheet_names, requested)


def _xlsx_sheet_has_data(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> bool:
    if (worksheet.max_row or 0) <= 1 or (worksheet.max_column or 0) == 0:
        return False
    header = next(worksheet.iter_rows(values_only=True, min_row=1, max_row=1), ())
    return any(value not in (None, "") for value in header)


def _xls_sheet_has_data(sheet: xlrd.sheet.Sheet) -> bool:
    if sheet.nrows <= 1 or sheet.ncols == 0:
        return False
    header = sheet.row_values(0)
    return any(value not in (None, "") for value in header)


def _build_row_signature(row: dict[str, Any]) -> str:
    normalized = {
        key: normalize_match_value(value)
        for key, value in sorted(row.items(), key=lambda item: item[0])
    }
    payload = json.dumps(normalized, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(payload.encode("utf-8")).hexdigest()


def _estimate_output_rows(
    config: ExportConfig,
    right_unique_counts: dict[tuple[str, ...], int],
    is_cancelled: Callable[[], bool] | None,
    *,
    dedupe_left_by_key: bool,
) -> int:
    unique_left_keys: set[tuple[str, ...]] = set()
    estimated_rows = 0

    left_stream = open_row_stream(config.left_path, config.left_sheet, config.encoding)
    try:
        for row in left_stream.iterator:
            _ensure_not_cancelled(is_cancelled)
            key = tuple(normalize_match_value(row.get(pair.left_raw)) for pair in config.match_pairs)
            if dedupe_left_by_key:
                if key in unique_left_keys:
                    continue
                unique_left_keys.add(key)
            estimated_rows += right_unique_counts.get(key, 0)
    finally:
        left_stream.close()

    return estimated_rows


def _open_xls_workbook(file_path: str) -> xlrd.book.Book:
    for candidate in ("gb18030", "gbk", "gb2312", "utf-8"):
        try:
            return xlrd.open_workbook(file_path, encoding_override=candidate)
        except Exception:
            continue
    return xlrd.open_workbook(file_path)


def _emit_progress(callback: Callable[[int, str], None] | None, percent: int, message: str) -> None:
    if callback is not None:
        callback(max(0, min(100, percent)), message)


def _ensure_not_cancelled(is_cancelled: Callable[[], bool] | None) -> None:
    if is_cancelled is not None and is_cancelled():
        raise ExportCancelled("导出已取消。")


def _create_lookup_backend(total_rows: int) -> "_LookupBackend":
    if total_rows <= LOOKUP_IN_MEMORY_MAX_ROWS:
        return _InMemoryLookup()
    return _SQLiteLookup()


class _LookupBackend:
    def add(
        self,
        key: tuple[str, ...],
        payload: dict[str, Any],
        *,
        row_signature: str,
        dedupe_by_key: bool,
        dedupe_by_row: bool,
    ) -> bool:
        raise NotImplementedError

    def find(self, key: tuple[str, ...]) -> list[dict[str, Any]]:
        raise NotImplementedError

    def close(self) -> None:
        raise NotImplementedError


class _InMemoryLookup(_LookupBackend):
    def __init__(self) -> None:
        self._data: dict[tuple[str, ...], list[dict[str, Any]]] = {}
        self._signatures: dict[tuple[str, ...], set[str]] = {}

    def add(
        self,
        key: tuple[str, ...],
        payload: dict[str, Any],
        *,
        row_signature: str,
        dedupe_by_key: bool,
        dedupe_by_row: bool,
    ) -> bool:
        if dedupe_by_key and key in self._data:
            return False
        if dedupe_by_row:
            signatures = self._signatures.setdefault(key, set())
            if row_signature in signatures:
                return False
            signatures.add(row_signature)
        self._data.setdefault(key, []).append(payload)
        return True

    def find(self, key: tuple[str, ...]) -> list[dict[str, Any]]:
        return self._data.get(key, [])

    def close(self) -> None:
        self._data.clear()
        self._signatures.clear()


class _SQLiteLookup(_LookupBackend):
    def __init__(self) -> None:
        self._temp_dir = tempfile.TemporaryDirectory(prefix="excel_tool_lookup_")
        self._db_path = os.path.join(self._temp_dir.name, "lookup.sqlite3")
        self._connection = sqlite3.connect(self._db_path)
        self._cursor = self._connection.cursor()
        self._initialized = False
        self._key_length = 0
        self._unique_mode = ""

    def add(
        self,
        key: tuple[str, ...],
        payload: dict[str, Any],
        *,
        row_signature: str,
        dedupe_by_key: bool,
        dedupe_by_row: bool,
    ) -> bool:
        if not self._initialized:
            self._initialize(len(key), dedupe_by_key=dedupe_by_key, dedupe_by_row=dedupe_by_row)
        payload_json = json.dumps(payload, ensure_ascii=False, default=str, sort_keys=True)
        row_signature_value = row_signature if dedupe_by_row else ""
        placeholders = ", ".join("?" for _ in range(self._key_length + 2))
        self._cursor.execute(
            f"INSERT OR IGNORE INTO lookup VALUES ({placeholders})",
            [*key, row_signature_value, payload_json],
        )
        return self._cursor.rowcount > 0

    def find(self, key: tuple[str, ...]) -> list[dict[str, Any]]:
        if not self._initialized:
            return []
        where_clause = " AND ".join(f"k{index} = ?" for index in range(self._key_length))
        query = f"SELECT payload FROM lookup WHERE {where_clause}"
        rows = self._connection.execute(query, key).fetchall()
        return [json.loads(item[0]) for item in rows]

    def close(self) -> None:
        if self._initialized:
            self._connection.commit()
        self._connection.close()
        self._temp_dir.cleanup()

    def _initialize(self, key_length: int, *, dedupe_by_key: bool, dedupe_by_row: bool) -> None:
        self._key_length = key_length
        key_columns = ", ".join(f"k{index} TEXT" for index in range(key_length))
        column_defs = [key_columns] if key_columns else []
        column_defs.append("row_signature TEXT")
        column_defs.append("payload TEXT")
        if dedupe_by_key:
            unique_columns = ", ".join(f"k{index}" for index in range(key_length))
            self._unique_mode = DEDUP_STRATEGY_BOTH_KEYS
        elif dedupe_by_row:
            unique_columns = ", ".join([*(f"k{index}" for index in range(key_length)), "row_signature"])
            self._unique_mode = DEDUP_STRATEGY_LEFT_KEY_RIGHT_ROW
        else:
            unique_columns = ""
            self._unique_mode = DEDUP_STRATEGY_NONE
        if unique_columns:
            create_sql = f"CREATE TABLE lookup ({', '.join(column_defs)}, UNIQUE ({unique_columns}))"
        else:
            create_sql = f"CREATE TABLE lookup ({', '.join(column_defs)})"
        self._cursor.execute(
            create_sql
        )
        if key_length:
            index_columns = ", ".join(f"k{index}" for index in range(key_length))
            self._cursor.execute(f"CREATE INDEX lookup_idx ON lookup ({index_columns})")
        self._initialized = True
