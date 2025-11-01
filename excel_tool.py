import sys
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QLabel, QLineEdit, QPushButton,
                             QGroupBox, QTableWidget, QTableWidgetItem, QFileDialog,
                             QComboBox, QCheckBox, QProgressDialog, QMessageBox,
                             QScrollArea, QSpinBox, QRadioButton, QButtonGroup)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
import pandas as pd
import numpy as np
import os
import xlrd
import datetime
import multiprocessing as mp
from concurrent.futures import ProcessPoolExecutor
import psutil
import gc
import warnings
warnings.filterwarnings('ignore')

class ExcelCompareTool(QMainWindow):
    """
    Excel文件比较工具主窗口类

    功能:
    - 提供两个Excel文件的比较界面
    - 支持列匹配和结果导出
    """

    def __init__(self):
        super().__init__()
        self._init_data_structures()
        self._init_ui()

    def _init_ui(self):
        """初始化用户界面"""
        self.setWindowTitle("Excel文件处理工具")
        self.setGeometry(100, 100, 1000, 600)
        self.setStyleSheet("QMainWindow { background-color: #f5f5f5; }")

        # 创建主布局
        self._setup_main_layout()

        # 添加文件选择区域
        self._add_file_selection_section()

    def _setup_main_layout(self):
        """设置主布局"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.main_layout = QVBoxLayout(central_widget)
        self.main_layout.setSpacing(15)
        self.main_layout.setContentsMargins(15, 15, 15, 15)

    def _add_file_selection_section(self):
        """添加文件选择区域"""
        # 初始化文件路径输入框
        self.file1_path = QLineEdit()
        self.file2_path = QLineEdit()
        self.file1_path.setStyleSheet("QLineEdit { padding: 5px; border: 1px solid #ccc; border-radius: 4px; }")
        self.file2_path.setStyleSheet("QLineEdit { padding: 5px; border: 1px solid #ccc; border-radius: 4px; }")

        # 创建文件选择组件
        file1_group = self._create_file_group("匹配源文件", self.file1_path, 1)
        file2_group = self._create_file_group("被匹配文件", self.file2_path, 2)

        # 创建文件选择组
        file_selection_group = QGroupBox("文件选择")
        file_selection_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #ddd;
                border-radius: 4px;
                margin-top: 5px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)

        # 添加编码选择区域
        encoding_group = self._create_encoding_selection()

        # 添加到主布局
        file_selection_layout = QHBoxLayout()
        file_selection_layout.setSpacing(15)
        file_selection_layout.setContentsMargins(10, 15, 10, 10)
        file_selection_layout.addWidget(file1_group)
        file_selection_layout.addWidget(file2_group)
        file_selection_layout.addWidget(encoding_group)
        file_selection_group.setLayout(file_selection_layout)
        self.main_layout.addWidget(file_selection_group)

        # 添加预览区域
        self._add_preview_section()

        # 添加列匹配设置区域
        self._add_column_matching_section()

        # 添加性能设置区域
        self._add_performance_settings_section()

        # 添加输出结果设置区域
        self._add_output_settings_section()

    def _init_data_structures(self):
        """初始化数据结构和默认值"""
        self.match_rows = []
        self.file1_checkboxes = []
        self.file2_checkboxes = []
        self.file1_path = None
        self.file2_path = None
        self.sheet1_combo = None
        self.sheet2_combo = None
        self.preview1 = None
        self.preview2 = None
        self.main_layout = None
        self.match_rows_widget = None
        self.match_rows_layout = None
        self.file1_checkbox_group = None
        self.file2_checkbox_group = None
        self.file1_checkbox_layout = None
        self.file2_checkbox_layout = None
        self.output_filename = None
        self.chunk_size_input = None
        self.process_count_input = None
        # 编码选择相关
        self.encoding_group = None
        self.gbk_radio = None
        self.utf8_radio = None

    def _add_preview_section(self):
        """添加数据预览区域"""
        preview_group = QGroupBox("数据预览")
        preview_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #ddd;
                border-radius: 4px;
                margin-top: 5px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)
        preview_layout = QHBoxLayout()
        preview_layout.setSpacing(2)
        preview_layout.setContentsMargins(10, 15, 10, 10)

        # 创建预览表格
        self.preview1_label = QLabel("未选择文件")
        self.preview2_label = QLabel("未选择文件")
        self.preview1 = self._create_preview_table()
        self.preview2 = self._create_preview_table()

        # 设置标签样式
        self._setup_preview_labels()

        # 创建并添加预览布局
        preview1_layout = self._create_preview_layout(self.preview1_label, self.preview1)
        preview2_layout = self._create_preview_layout(self.preview2_label, self.preview2)

        preview_layout.addLayout(preview1_layout)
        preview_layout.addLayout(preview2_layout)
        preview_group.setLayout(preview_layout)
        self.main_layout.addWidget(preview_group)

    def _create_preview_table(self):
        """创建预览表格"""
        table = QTableWidget()
        table.setFixedHeight(250)
        table.setFixedWidth(480)
        table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        table.setContentsMargins(0, 0, 0, 0)
        table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                border: 1px solid #ddd;
                border-radius: 4px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 5px;
                border: none;
            }
        """)
        return table

    def _setup_preview_labels(self):
        """设置预览标签样式"""
        self.preview1_label.setContentsMargins(0, 0, 0, 5)
        self.preview2_label.setContentsMargins(0, 0, 0, 5)
        self.preview1_label.setStyleSheet("QLabel { font-weight: bold; color: #333; }")
        self.preview2_label.setStyleSheet("QLabel { font-weight: bold; color: #333; }")

    def _create_preview_layout(self, label, table):
        """创建单个预览布局"""
        layout = QVBoxLayout()
        layout.setSpacing(0)
        layout.addWidget(label)
        layout.addWidget(table)
        return layout

    def _add_column_matching_section(self):
        """添加列匹配设置区域"""
        match_group = QGroupBox("列匹配设置")
        match_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #ddd;
                border-radius: 4px;
                margin-top: 5px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)
        match_layout = QVBoxLayout()
        match_layout.setSpacing(5)
        match_layout.setContentsMargins(10, 15, 10, 10)

        # 添加按钮
        self._add_matching_buttons(match_layout)

        # 添加匹配行容器
        self._add_match_rows_container(match_layout)

        match_group.setLayout(match_layout)
        self.main_layout.addWidget(match_group)

    def _add_performance_settings_section(self):
        """添加性能设置区域"""
        perf_group = QGroupBox("性能设置")
        perf_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #ddd;
                border-radius: 4px;
                margin-top: 5px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)
        perf_layout = QHBoxLayout()
        perf_layout.setSpacing(20)
        perf_layout.setContentsMargins(10, 15, 10, 10)

        # 分块大小设置
        chunk_layout = QHBoxLayout()
        chunk_layout.addWidget(QLabel("分块大小:"))
        self.chunk_size_input = QSpinBox()
        self.chunk_size_input.setRange(1000, 100000)
        self.chunk_size_input.setValue(10000)
        self.chunk_size_input.setSuffix(" 行")
        chunk_layout.addWidget(self.chunk_size_input)

        # 进程数设置
        process_layout = QHBoxLayout()
        process_layout.addWidget(QLabel("并行进程数:"))
        self.process_count_input = QSpinBox()
        cpu_count = mp.cpu_count()
        self.process_count_input.setRange(1, cpu_count)
        self.process_count_input.setValue(min(4, cpu_count))  # 默认使用4个进程或CPU核心数
        process_layout.addWidget(self.process_count_input)

        # 系统信息显示
        info_layout = QVBoxLayout()
        cpu_info = QLabel(f"CPU核心数: {cpu_count}")
        memory_info = QLabel(f"可用内存: {psutil.virtual_memory().available // (1024**3):.1f} GB")
        cpu_info.setStyleSheet("QLabel { color: #666; font-size: 12px; }")
        memory_info.setStyleSheet("QLabel { color: #666; font-size: 12px; }")
        info_layout.addWidget(cpu_info)
        info_layout.addWidget(memory_info)

        perf_layout.addLayout(chunk_layout)
        perf_layout.addLayout(process_layout)
        perf_layout.addLayout(info_layout)
        perf_layout.addStretch()

        perf_group.setLayout(perf_layout)
        self.main_layout.addWidget(perf_group)

    def _handle_add_match_row(self):
        """
        处理增加匹配行按钮点击事件
        """
        self.add_match_row()

    def _handle_delete_match_row(self):
        """
        处理删除匹配行按钮点击事件
        """
        self.delete_match_row()

    def _add_matching_buttons(self, parent_layout):
        """添加匹配操作的按钮"""
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)

        add_button = QPushButton("增加")
        add_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-weight: bold;
                padding: 5px 10px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        add_button.clicked.connect(self._handle_add_match_row)

        delete_button = QPushButton("删除")
        delete_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-weight: bold;
                padding: 5px 10px;
                border: none;
                border-radius: 4px;
            }
            QPushButton:hover {
                background-color: #d32f2f;
            }
        """)
        delete_button.clicked.connect(self._handle_delete_match_row)

        button_layout.addWidget(add_button)
        button_layout.addWidget(delete_button)
        button_layout.addStretch()

        parent_layout.addLayout(button_layout)

    def _add_match_rows_container(self, parent_layout):
        """添加匹配行容器"""
        self.match_rows_widget = QWidget()
        self.match_rows_layout = QVBoxLayout(self.match_rows_widget)
        self.match_rows_layout.setSpacing(5)
        self.match_rows_layout.setContentsMargins(0, 0, 0, 0)
        parent_layout.addWidget(self.match_rows_widget)

    def _add_output_settings_section(self):
        """添加输出结果设置区域"""
        output_group = QGroupBox("输出结果设置")
        output_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #ddd;
                border-radius: 4px;
                margin-top: 5px;
                padding-top: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 3px;
            }
        """)
        output_layout = QVBoxLayout()
        output_layout.setSpacing(10)
        output_layout.setContentsMargins(10, 15, 10, 10)

        # 添加多选框区域
        self._add_output_checkboxes(output_layout)

        # 创建文件名和按钮的布局
        filename_button_layout = QHBoxLayout()

        # 添加输出文件名设置
        self._add_output_filename_setting(filename_button_layout)

        # 添加导出按钮
        self._add_export_button(filename_button_layout)

        # 添加布局到主输出布局
        output_layout.addLayout(filename_button_layout)

        output_group.setLayout(output_layout)
        self.main_layout.addWidget(output_group)

    def _add_output_checkboxes(self, parent_layout):
        """添加输出列多选框"""
        checkboxes_layout = QHBoxLayout()

        # 文件1的多选框组
        self.file1_checkbox_group = self._create_checkbox_group("匹配源文件输出列")
        self.file1_checkbox_group.setStyleSheet("QGroupBox { border: none; }")

        # 文件2的多选框组
        self.file2_checkbox_group = self._create_checkbox_group("被匹配文件输出列")
        self.file2_checkbox_group.setStyleSheet("QGroupBox { border: none; }")

        checkboxes_layout.addWidget(self.file1_checkbox_group)
        checkboxes_layout.addWidget(self.file2_checkbox_group)
        parent_layout.addLayout(checkboxes_layout)

    def _create_checkbox_group(self, title):
        """创建多选框组"""
        group = QGroupBox(title)
        container = QWidget()

        if title == "匹配源文件输出列":
            self.file1_checkbox_layout = QVBoxLayout(container)
            self.file1_checkbox_layout.setSpacing(10)
            self.file1_checkbox_layout.setContentsMargins(10, 10, 10, 10)
        else:
            self.file2_checkbox_layout = QVBoxLayout(container)
            self.file2_checkbox_layout.setSpacing(10)
            self.file2_checkbox_layout.setContentsMargins(10, 10, 10, 10)

        scroll = QScrollArea()
        scroll.setWidget(container)
        scroll.setWidgetResizable(True)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setMinimumHeight(150)
        scroll.setMaximumHeight(300)

        group_layout = QVBoxLayout(group)
        group_layout.setSpacing(10)
        group_layout.setContentsMargins(5, 5, 5, 5)
        group_layout.addWidget(scroll)

        return group

    def _add_output_filename_setting(self, parent_layout):
        """添加输出文件名设置"""
        filename_layout = QHBoxLayout()
        filename_layout.addWidget(QLabel("输出文件名:"))
        self.output_filename = QLineEdit("匹配结果")
        filename_layout.addWidget(self.output_filename)
        filename_layout.addWidget(QLabel(".xlsx"))
        parent_layout.addLayout(filename_layout)

    def _add_export_button(self, parent_layout):
        """添加导出按钮"""
        export_button = QPushButton("导出结果")
        export_button.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
                border: none;
                border-radius: 4px;
                min-width: 100px;
            }
            QPushButton:hover {
                background-color: #0b7dda;
            }
        """)
        export_button.clicked.connect(self._handle_export_result)
        parent_layout.addWidget(export_button)
    def _handle_export_result(self):
        """
        处理导出结果按钮点击事件
        """
        self.export_result()

    def _handle_file_selection(self, file_path, file_num):
        """
        处理文件选择后的通用逻辑

        参数:
            file_path: 选择的文件路径
            file_num: 文件编号(1或2)
        """
        if file_num == 1:
            self.file1_path.setText(file_path)
            self.preview1_label.setText(f"匹配源文件: {file_path.split('/')[-1]}")
            self.show_preview(file_path, self.preview1)
        else:
            self.file2_path.setText(file_path)
            self.preview2_label.setText(f"被匹配文件: {file_path.split('/')[-1]}")
            self.show_preview(file_path, self.preview2)

    def _select_file(self, file_num):
        """
        打开文件选择对话框并处理选择结果

        参数:
            file_num: 文件编号(1或2)
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"选择文件{file_num}",
            "",
            "Excel/CSV files (*.xlsx *.xls *.csv);;Excel files (*.xlsx *.xls);;CSV files (*.csv);;All files (*.*)"
        )

        if file_path:
            self._handle_file_selection(file_path, file_num)

    def _handle_sheet_change(self, file_path, sheet_name, preview_table):
        """
        处理sheet变更后的通用逻辑

        参数:
            file_path: 文件路径
            sheet_name: 选择的sheet名称
            preview_table: 要更新的预览表格
        """
        if file_path and sheet_name:
            self.show_preview(file_path, preview_table, sheet_name)

    def _on_sheet_changed(self, file_num):
        """
        处理sheet选择变化事件

        参数:
            file_num: 文件编号(1或2)
        """
        file_path = self.file1_path.text() if file_num == 1 else self.file2_path.text()
        sheet_combo = self.sheet1_combo if file_num == 1 else self.sheet2_combo
        preview_table = self.preview1 if file_num == 1 else self.preview2

        self._handle_sheet_change(file_path, sheet_combo.currentText(), preview_table)

    def _create_file_group(self, title, line_edit, file_num):
        """
        创建文件选择组

        参数:
            title: 组标题
            line_edit: 文件路径输入框
            file_num: 文件编号(1或2)

        返回:
            QGroupBox: 包含文件选择控件的组
        """
        group = QGroupBox(title)
        layout = QHBoxLayout()

        # 设置文件路径输入框
        line_edit.setReadOnly(True)
        layout.addWidget(line_edit)

        # 添加文件选择按钮
        select_button = QPushButton("选择文件")
        select_button.clicked.connect(lambda: self._select_file(file_num))
        layout.addWidget(select_button)

        # 添加sheet选择下拉框
        sheet_combo = self._create_sheet_combo(file_num)
        layout.addWidget(sheet_combo)

        group.setLayout(layout)
        return group

    def _create_sheet_combo(self, file_num):
        """创建sheet选择下拉框"""
        combo = QComboBox()
        combo.setEnabled(False)
        combo.currentIndexChanged.connect(lambda: self._on_sheet_changed(file_num))

        # 保存引用
        if file_num == 1:
            self.sheet1_combo = combo
        else:
            self.sheet2_combo = combo

        return combo

    def _create_encoding_selection(self):
        """创建编码选择组"""
        group = QGroupBox("文件编码")
        layout = QVBoxLayout()
        layout.setSpacing(5)
        layout.setContentsMargins(10, 10, 10, 10)

        # 创建单选按钮组
        self.encoding_group = QButtonGroup()

        # GBK编码单选按钮（默认选中）
        self.gbk_radio = QRadioButton("GBK")
        self.gbk_radio.setChecked(True)
        self.gbk_radio.toggled.connect(self._on_encoding_changed)

        # UTF-8编码单选按钮
        self.utf8_radio = QRadioButton("UTF-8")
        self.utf8_radio.toggled.connect(self._on_encoding_changed)

        # 添加到按钮组
        self.encoding_group.addButton(self.gbk_radio, 0)
        self.encoding_group.addButton(self.utf8_radio, 1)

        # 添加到布局
        layout.addWidget(self.gbk_radio)
        layout.addWidget(self.utf8_radio)
        layout.addStretch()

        group.setLayout(layout)
        group.setFixedWidth(120)
        return group

    def _on_encoding_changed(self):
        """编码选择变化时的处理"""
        # 如果已经选择了文件，重新加载预览
        if self.file1_path and self.file1_path.text():
            self.show_preview(self.file1_path.text(), self.preview1)
        if self.file2_path and self.file2_path.text():
            self.show_preview(self.file2_path.text(), self.preview2)

    def get_selected_encoding(self):
        """获取当前选择的编码"""
        # 如果编码选择组件还没有初始化，返回默认编码
        if not hasattr(self, 'gbk_radio') or self.gbk_radio is None:
            return 'gbk'
        return 'gbk' if self.gbk_radio.isChecked() else 'utf-8'

    def select_file(self, file_num):
        """
        公开的文件选择方法

        参数:
            file_num: 文件编号(1或2)
        """
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"选择文件{file_num}",
            "",
            "Excel/CSV files (*.xlsx *.xls *.csv);;Excel files (*.xlsx *.xls);;CSV files (*.csv);;All files (*.*)"
        )

        if file_path:
            self._handle_file_selection(file_path, file_num)

    def on_sheet_changed(self, file_num):
        """
        公开的sheet变更处理方法

        参数:
            file_num: 文件编号(1或2)
        """
        file_path = self.file1_path.text() if file_num == 1 else self.file2_path.text()
        sheet_combo = self.sheet1_combo if file_num == 1 else self.sheet2_combo
        preview_table = self.preview1 if file_num == 1 else self.preview2

        self._handle_sheet_change(file_path, sheet_combo.currentText(), preview_table)

    def show_preview(self, file_path, preview_table, sheet_name=None):
        try:
            # 检查文件类型
            file_ext = os.path.splitext(file_path)[1].lower()

            # 获取当前选择的编码
            encoding = self.get_selected_encoding()

            if file_ext == '.csv':
                # 处理CSV文件
                self._handle_csv_preview(file_path, preview_table, encoding)
                return  # CSV处理完成后直接返回
            elif file_ext == '.xls':
                # 使用xlrd处理xls文件
                # 尝试使用常见的中文编码
                encodings = ['gb18030', 'gbk', 'gb2312', 'utf-8']
                workbook = None

                for encoding in encodings:
                    try:
                        workbook = xlrd.open_workbook(file_path, encoding_override=encoding)
                        # 尝试读取一个单元格的内容来验证编码是否正确
                        sheet = workbook.sheet_by_index(0)
                        test_value = str(sheet.cell_value(0, 0))
                        if not any(c == '?' or c == '�' or ord(c) > 0xffff for c in test_value):
                            break
                    except Exception:
                        continue

                if not workbook:
                    # 如果所有编码都失败，使用默认编码
                    workbook = xlrd.open_workbook(file_path)

                # 如果是新选择的文件，更新sheet列表
                if not sheet_name:
                    sheet_names = workbook.sheet_names()

                    # 更新对应的sheet下拉列表
                    sheet_combo = self.sheet1_combo if preview_table == self.preview1 else self.sheet2_combo
                    sheet_combo.clear()
                    sheet_combo.addItems(sheet_names)
                    sheet_combo.setEnabled(True)
                    sheet_name = sheet_names[0]  # 默认选择第一个sheet

                # 读取指定sheet的前5行
                sheet = workbook.sheet_by_name(sheet_name)

                # 创建DataFrame
                data = []
                headers = []

                # 获取表头
                for col in range(sheet.ncols):
                    headers.append(str(sheet.cell_value(0, col)))

                # 获取数据（前5行）
                for row in range(1, min(6, sheet.nrows)):
                    row_data = []
                    for col in range(sheet.ncols):
                        cell_value = sheet.cell_value(row, col)
                        # 处理日期类型
                        if sheet.cell_type(row, col) == xlrd.XL_CELL_DATE:
                            try:
                                cell_value = xlrd.xldate.xldate_as_datetime(cell_value, workbook.datemode).strftime('%Y-%m-%d')
                            except:
                                pass
                        row_data.append(cell_value)
                    data.append(row_data)

                # 创建DataFrame
                df = pd.DataFrame(data, columns=headers)
            else:
                # 使用openpyxl读取Excel文件
                # 如果是新选择的文件，更新sheet列表
                if not sheet_name:
                    # 使用openpyxl读取
                    excel_file = pd.ExcelFile(file_path, engine='openpyxl')
                    sheet_names = excel_file.sheet_names

                    # 更新对应的sheet下拉列表
                    sheet_combo = self.sheet1_combo if preview_table == self.preview1 else self.sheet2_combo
                    sheet_combo.clear()
                    sheet_combo.addItems(sheet_names)
                    sheet_combo.setEnabled(True)
                    sheet_name = sheet_names[0]  # 默认选择第一个sheet

                # 读取指定sheet的前5行
                df = pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name, nrows=5)

            # 设置表格的行数和列数
            preview_table.setRowCount(len(df.index))
            preview_table.setColumnCount(len(df.columns))

            # 设置表头（Excel列标识与列名拼接）
            excel_columns = [chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26) for i in range(len(df.columns))]
            header_labels = [f"{col}-{str(val)}" for col, val in zip(excel_columns, df.columns)]
            preview_table.setHorizontalHeaderLabels(header_labels)

            # 填充数据
            for i in range(len(df.index)):
                for j in range(len(df.columns)):
                    item = QTableWidgetItem(str(df.iloc[i, j]))
                    preview_table.setItem(i, j, item)

            # 调整列宽以适应内容
            preview_table.resizeColumnsToContents()
            preview_table.resizeRowsToContents()

            # 更新列匹配下拉列表和输出列多选框
            if preview_table == self.preview1:
                self.update_file1_columns(header_labels)
            else:
                self.update_file2_columns(header_labels)

        except Exception as e:
            # 清空表格并显示错误信息
            preview_table.setRowCount(0)
            preview_table.setColumnCount(1)
            preview_table.setHorizontalHeaderLabels(["错误"])
            preview_table.setRowCount(1)
            preview_table.setItem(0, 0, QTableWidgetItem(str(e)))

    def _handle_csv_preview(self, file_path, preview_table, encoding):
        """处理CSV文件预览"""
        try:
            # 读取CSV文件的前5行进行预览
            df = pd.read_csv(file_path, encoding=encoding, nrows=5)

            # 如果是新选择的文件，禁用sheet选择（CSV没有sheet概念）
            sheet_combo = self.sheet1_combo if preview_table == self.preview1 else self.sheet2_combo
            sheet_combo.clear()
            sheet_combo.addItem("CSV文件")
            sheet_combo.setEnabled(False)

            # 设置表格的行数和列数
            preview_table.setRowCount(len(df.index))
            preview_table.setColumnCount(len(df.columns))

            # 设置表头（Excel列标识与列名拼接）
            excel_columns = [chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26) for i in range(len(df.columns))]
            header_labels = [f"{col}-{str(val)}" for col, val in zip(excel_columns, df.columns)]
            preview_table.setHorizontalHeaderLabels(header_labels)

            # 填充数据
            for i in range(len(df.index)):
                for j in range(len(df.columns)):
                    item = QTableWidgetItem(str(df.iloc[i, j]))
                    preview_table.setItem(i, j, item)

            # 调整列宽以适应内容
            preview_table.resizeColumnsToContents()
            preview_table.resizeRowsToContents()

            # 更新列匹配下拉列表和输出列多选框
            if preview_table == self.preview1:
                self.update_file1_columns(header_labels)
            else:
                self.update_file2_columns(header_labels)

        except UnicodeDecodeError:
            # 如果编码错误，尝试另一种编码
            try:
                alternative_encoding = 'utf-8' if encoding == 'gbk' else 'gbk'
                df = pd.read_csv(file_path, encoding=alternative_encoding, nrows=5)

                # 显示编码警告
                QMessageBox.warning(self, "编码警告",
                                  f"使用{encoding}编码读取失败，已自动切换到{alternative_encoding}编码")

                # 更新编码选择
                if alternative_encoding == 'utf-8':
                    self.utf8_radio.setChecked(True)
                else:
                    self.gbk_radio.setChecked(True)

                # 处理成功读取的数据
                # 如果是新选择的文件，禁用sheet选择（CSV没有sheet概念）
                sheet_combo = self.sheet1_combo if preview_table == self.preview1 else self.sheet2_combo
                sheet_combo.clear()
                sheet_combo.addItem("CSV文件")
                sheet_combo.setEnabled(False)

                # 设置表格的行数和列数
                preview_table.setRowCount(len(df.index))
                preview_table.setColumnCount(len(df.columns))

                # 设置表头（Excel列标识与列名拼接）
                excel_columns = [chr(65 + i) if i < 26 else chr(64 + i//26) + chr(65 + i%26) for i in range(len(df.columns))]
                header_labels = [f"{col}-{str(val)}" for col, val in zip(excel_columns, df.columns)]
                preview_table.setHorizontalHeaderLabels(header_labels)

                # 填充数据
                for i in range(len(df.index)):
                    for j in range(len(df.columns)):
                        item = QTableWidgetItem(str(df.iloc[i, j]))
                        preview_table.setItem(i, j, item)

                # 调整列宽以适应内容
                preview_table.resizeColumnsToContents()
                preview_table.resizeRowsToContents()

                # 更新列匹配下拉列表和输出列多选框
                if preview_table == self.preview1:
                    self.update_file1_columns(header_labels)
                else:
                    self.update_file2_columns(header_labels)

            except Exception as e:
                # 如果两种编码都失败，显示错误
                preview_table.setRowCount(1)
                preview_table.setColumnCount(1)
                preview_table.setHorizontalHeaderLabels(["错误"])
                preview_table.setItem(0, 0, QTableWidgetItem(f"CSV文件编码错误: {str(e)}"))
        except Exception as e:
            # 其他错误
            preview_table.setRowCount(1)
            preview_table.setColumnCount(1)
            preview_table.setHorizontalHeaderLabels(["错误"])
            preview_table.setItem(0, 0, QTableWidgetItem(f"读取CSV文件出错: {str(e)}"))

    def update_file1_columns(self, columns):
        # 获取已选择的匹配列
        selected_columns1 = []
        for row in self.match_rows:
            text1 = row[0].currentText()
            if text1:
                selected_columns1.append(text1)

        # 如果没有匹配行，自动添加一个默认匹配行
        if not self.match_rows and columns:
            self.add_match_row()
            selected_columns1 = []

            # 自动设置第一列为默认匹配列
            if columns:
                self.match_rows[0][0].setCurrentText(columns[0])

        # 更新文件1的下拉列表选项
        for row in self.match_rows:
            row[0].clear()
            row[0].addItems(columns)

        # 更新文件1的输出列多选框（如果布局已经初始化）
        if hasattr(self, 'file1_checkbox_layout') and self.file1_checkbox_layout is not None:
            for checkbox in self.file1_checkboxes:
                self.file1_checkbox_layout.removeWidget(checkbox)
                checkbox.deleteLater()
            self.file1_checkboxes.clear()

            # 只添加未被选为匹配列的列到输出多选框
            for column in columns:
                if column not in selected_columns1:
                    checkbox = QCheckBox(column)
                    self.file1_checkboxes.append(checkbox)
                    self.file1_checkbox_layout.addWidget(checkbox)

    def update_file2_columns(self, columns):
        # 获取已选择的匹配列
        selected_columns2 = []
        for row in self.match_rows:
            text2 = row[1].currentText()
            if text2:
                selected_columns2.append(text2)

        # 如果没有匹配行，自动添加一个默认匹配行
        if not self.match_rows and columns:
            self.add_match_row()
            selected_columns2 = []

            # 自动设置第一列为默认匹配列
            if columns:
                self.match_rows[0][1].setCurrentText(columns[0])

        # 更新文件2的下拉列表选项
        for row in self.match_rows:
            row[1].clear()
            row[1].addItems(columns)

        # 更新文件2的输出列多选框（如果布局已经初始化）
        if hasattr(self, 'file2_checkbox_layout') and self.file2_checkbox_layout is not None:
            for checkbox in self.file2_checkboxes:
                self.file2_checkbox_layout.removeWidget(checkbox)
                checkbox.deleteLater()
            self.file2_checkboxes.clear()

            # 只添加未被选为匹配列的列到输出多选框
            for column in columns:
                if column not in selected_columns2:
                    checkbox = QCheckBox(column)
                    self.file2_checkboxes.append(checkbox)
                    self.file2_checkbox_layout.addWidget(checkbox)

    def add_match_row(self):
        # 创建一行匹配选择
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        combo1 = QComboBox()
        combo2 = QComboBox()

        # 获取已选择的列
        selected_columns1 = []
        selected_columns2 = []
        for row in self.match_rows:
            text1 = row[0].currentText()
            text2 = row[1].currentText()
            if text1:
                selected_columns1.append(text1)
            if text2:
                selected_columns2.append(text2)

        # 如果已经有文件1的数据，添加未被选择的列到下拉列表
        if self.preview1.columnCount() > 0:
            headers = [self.preview1.horizontalHeaderItem(i).text()
                      for i in range(self.preview1.columnCount())]
            available_headers = [h for h in headers if h not in selected_columns1]
            combo1.addItems(available_headers)
            if available_headers:  # 如果有可用的列，设置第一个为默认值
                combo1.setCurrentText(available_headers[0])

        # 如果已经有文件2的数据，添加未被选择的列到下拉列表
        if self.preview2.columnCount() > 0:
            headers = [self.preview2.horizontalHeaderItem(i).text()
                      for i in range(self.preview2.columnCount())]
            available_headers = [h for h in headers if h not in selected_columns2]
            combo2.addItems(available_headers)
            if available_headers:  # 如果有可用的列，设置第一个为默认值
                combo2.setCurrentText(available_headers[0])

        # 添加选择变化事件处理
        combo1.currentIndexChanged.connect(lambda index: self.on_match_column_changed())
        combo2.currentIndexChanged.connect(lambda index: self.on_match_column_changed())

        row_layout.addWidget(combo1)
        row_layout.addWidget(combo2)

        # 将下拉列表添加到匹配行列表中
        self.match_rows.append([combo1, combo2])

        # 将行布局添加到匹配行容器中
        self.match_rows_layout.addWidget(row_widget)

        # 更新所有下拉列表的可选项
        self.update_available_columns()

        # 手动触发更新输出列多选框
        self.on_match_column_changed()

    def on_match_column_changed(self):
        # 更新所有下拉列表的可选项
        self.update_available_columns()

        # 获取当前所有已选择的匹配列
        selected_columns1 = []
        selected_columns2 = []
        for row in self.match_rows:
            text1 = row[0].currentText()
            text2 = row[1].currentText()
            if text1:
                selected_columns1.append(text1)
            if text2:
                selected_columns2.append(text2)

        # 更新文件1的输出列多选框（如果布局已经初始化）
        if hasattr(self, 'file1_checkbox_layout') and self.file1_checkbox_layout is not None:
            for checkbox in self.file1_checkboxes:
                self.file1_checkbox_layout.removeWidget(checkbox)
                checkbox.deleteLater()
            self.file1_checkboxes.clear()

            # 获取文件1的所有列
            if self.preview1.columnCount() > 0:
                columns1 = [self.preview1.horizontalHeaderItem(i).text()
                           for i in range(self.preview1.columnCount())]
                # 只添加未被选为匹配列的列到输出多选框
                for column in columns1:
                    if column not in selected_columns1:
                        checkbox = QCheckBox(column)
                        self.file1_checkboxes.append(checkbox)
                        self.file1_checkbox_layout.addWidget(checkbox)

        # 更新文件2的输出列多选框（如果布局已经初始化）
        if hasattr(self, 'file2_checkbox_layout') and self.file2_checkbox_layout is not None:
            for checkbox in self.file2_checkboxes:
                self.file2_checkbox_layout.removeWidget(checkbox)
                checkbox.deleteLater()
            self.file2_checkboxes.clear()

            # 获取文件2的所有列
            if self.preview2.columnCount() > 0:
                columns2 = [self.preview2.horizontalHeaderItem(i).text()
                           for i in range(self.preview2.columnCount())]
                # 只添加未被选为匹配列的列到输出多选框
                for column in columns2:
                    if column not in selected_columns2:
                        checkbox = QCheckBox(column)
                        self.file2_checkboxes.append(checkbox)
                        self.file2_checkbox_layout.addWidget(checkbox)

    def update_available_columns(self):
        # 获取所有可用列
        all_columns1 = []
        all_columns2 = []

        if self.preview1.columnCount() > 0:
            all_columns1 = [self.preview1.horizontalHeaderItem(i).text()
                           for i in range(self.preview1.columnCount())]
        if self.preview2.columnCount() > 0:
            all_columns2 = [self.preview2.horizontalHeaderItem(i).text()
                           for i in range(self.preview2.columnCount())]

        # 获取所有已选择的列
        selected_columns1 = []
        selected_columns2 = []

        # 先收集所有已选择的列
        for combo1, combo2 in self.match_rows:
            text1 = combo1.currentText()
            text2 = combo2.currentText()
            if text1:
                selected_columns1.append(text1)
            if text2:
                selected_columns2.append(text2)

        # 更新每个下拉列表
        for combo1, combo2 in self.match_rows:
            current1 = combo1.currentText()
            current2 = combo2.currentText()

            # 暂时断开信号连接
            combo1.blockSignals(True)
            combo2.blockSignals(True)

            # 更新文件1的下拉列表
            combo1.clear()
            available_columns1 = [col for col in all_columns1 if col not in selected_columns1 or col == current1]
            combo1.addItems(available_columns1)
            if current1 in available_columns1:
                combo1.setCurrentText(current1)

            # 更新文件2的下拉列表
            combo2.clear()
            available_columns2 = [col for col in all_columns2 if col not in selected_columns2 or col == current2]
            combo2.addItems(available_columns2)
            if current2 in available_columns2:
                combo2.setCurrentText(current2)

            # 恢复信号连接
            combo1.blockSignals(False)
            combo2.blockSignals(False)

    def delete_match_row(self):
        if self.match_rows:
            # 获取最后一行的组件
            last_row = self.match_rows.pop()
            layout = last_row[0].parent().layout()

            # 删除组件
            for combo in last_row:
                layout.removeWidget(combo)
                combo.deleteLater()

            # 删除布局
            self.match_rows_layout.removeItem(layout)
            layout.deleteLater()

            # 更新所有下拉列表的可选项
            self.update_available_columns()

    def on_export_finished(self, output_path, loading):
        loading.close()
        QMessageBox.information(self, "导出成功", f"数据已成功匹配并导出到：\n{output_path}")

    def on_export_error(self, error_msg, loading):
        loading.close()
        QMessageBox.critical(self, "导出错误", f"导出过程中发生错误：\n{error_msg}")

    def on_progress_updated(self, progress, status, loading):
        """更新进度条和状态信息"""
        if hasattr(loading, 'setValue'):
            loading.setValue(progress)
        loading.setLabelText(status)

    def export_result(self):
        if not self.file1_path.text() or not self.file2_path.text():
            return

        # 创建带动画的进度对话框
        loading = QProgressDialog(self)
        loading.setWindowTitle("处理中")
        loading.setLabelText("正在匹配数据，请稍候...")
        loading.setRange(0, 100)  # 设置为0-100的进度条
        loading.setValue(0)
        loading.setWindowModality(Qt.WindowModality.WindowModal)
        loading.setMinimumDuration(0)  # 立即显示
        loading.setCancelButton(None)  # 移除取消按钮
        loading.setStyleSheet("""
            QProgressDialog {
                background-color: #f5f5f5;
                border-radius: 8px;
                padding: 20px;
            }
            QProgressDialog QLabel {
                color: #333333;
                font-size: 14px;
                margin-bottom: 10px;
            }
            QProgressBar {
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                text-align: center;
                background-color: #ffffff;
            }
            QProgressBar::chunk {
                background-color: #4a90e2;
                border-radius: 3px;
            }
        """)
        loading.show()

        # 获取匹配列
        match_columns = []
        for combo1, combo2 in self.match_rows:
            if combo1.currentText() and combo2.currentText():
                match_columns.append((combo1.currentText(), combo2.currentText()))

        if not match_columns:
            loading.close()
            return

        # 获取选中的输出列
        output_columns = []
        for checkbox in self.file1_checkboxes:
            if checkbox.isChecked():
                output_columns.append(checkbox.text().split('-')[1])
        for checkbox in self.file2_checkboxes:
            if checkbox.isChecked():
                col_name = checkbox.text().split('-')[1]
                if col_name not in output_columns:  # 避免重复列
                    output_columns.append(col_name)

        # 创建工作线程
        self.worker = ExportWorker(
            self.file1_path.text(),
            self.file2_path.text(),
            self.sheet1_combo.currentText(),
            self.sheet2_combo.currentText(),
            match_columns,
            output_columns,
            self.output_filename.text(),
            self.chunk_size_input.value(),
            self.process_count_input.value(),
            self.get_selected_encoding()
        )

        # 连接信号
        self.worker.finished.connect(lambda output_path: self.on_export_finished(output_path, loading))
        self.worker.error.connect(lambda error_msg: self.on_export_error(error_msg, loading))
        self.worker.progress_updated.connect(lambda progress, status: self.on_progress_updated(progress, status, loading))

        # 启动工作线程
        self.worker.start()


class ExportWorker(QThread):
    finished = pyqtSignal(str)
    error = pyqtSignal(str)
    progress_updated = pyqtSignal(int, str)  # 进度百分比和状态信息

    def __init__(self, file1_path, file2_path, sheet1_name, sheet2_name,
                 match_columns, output_columns, output_filename, chunk_size=10000, process_count=4, encoding='gbk'):
        super().__init__()
        self.file1_path = file1_path
        self.file2_path = file2_path
        self.sheet1_name = sheet1_name
        self.sheet2_name = sheet2_name
        self.match_columns = match_columns
        self.output_columns = output_columns
        self.output_filename = output_filename
        self.chunk_size = chunk_size
        self.process_count = process_count
        self.encoding = encoding

    def run(self):
        try:
            self.progress_updated.emit(5, "正在分析文件结构...")

            # 获取文件扩展名
            file1_ext = os.path.splitext(self.file1_path)[1].lower()
            file2_ext = os.path.splitext(self.file2_path)[1].lower()

            # 获取文件行数以确定分块策略
            total_rows1 = self._get_file_row_count(self.file1_path, self.sheet1_name, file1_ext)
            total_rows2 = self._get_file_row_count(self.file2_path, self.sheet2_name, file2_ext)

            self.progress_updated.emit(10, f"文件1: {total_rows1}行, 文件2: {total_rows2}行")

            # 如果数据量较小，使用传统方法
            if total_rows1 < 50000 and total_rows2 < 50000:
                return self._process_small_files()

            # 大文件使用分块并行处理
            return self._process_large_files_parallel(total_rows1, total_rows2, file1_ext, file2_ext)

        except Exception as e:
            self.error.emit(str(e))

    def _get_file_row_count(self, file_path, sheet_name, file_ext):
        """获取文件行数"""
        try:
            if file_ext == '.csv':
                # 对于CSV文件，快速计算行数
                with open(file_path, 'r', encoding=self.encoding) as f:
                    return sum(1 for _ in f)
            elif file_ext == '.xls':
                workbook = xlrd.open_workbook(file_path)
                sheet = workbook.sheet_by_name(sheet_name)
                return sheet.nrows
            else:
                # 对于xlsx文件，使用openpyxl快速获取行数
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb[sheet_name]
                    return ws.max_row
                except ImportError:
                    # 如果openpyxl不可用，使用pandas
                    df = pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name, nrows=1)
                    return len(pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name))
        except Exception:
            return 0

    def _process_small_files(self):
        """处理小文件的传统方法"""
        self.progress_updated.emit(20, "使用传统方法处理小文件...")

        # 读取完整文件
        df1 = self._read_complete_file(self.file1_path, self.sheet1_name,
                                     os.path.splitext(self.file1_path)[1].lower())
        df2 = self._read_complete_file(self.file2_path, self.sheet2_name,
                                     os.path.splitext(self.file2_path)[1].lower())

        self.progress_updated.emit(60, "正在进行数据匹配...")

        # 执行匹配
        result = self._perform_merge(df1, df2)

        self.progress_updated.emit(90, "正在导出结果...")

        # 导出结果
        output_path = self._export_result(result)
        self.finished.emit(output_path)

    def _process_large_files_parallel(self, total_rows1, total_rows2, file1_ext, file2_ext):
        """使用并行处理大文件"""
        self.progress_updated.emit(15, "正在准备并行处理...")

        # 读取文件2到内存（通常被匹配文件较小）
        self.progress_updated.emit(20, "正在加载被匹配文件...")
        df2 = self._read_complete_file(self.file2_path, self.sheet2_name, file2_ext)

        # 优化df2的内存使用
        df2 = optimize_dataframe_memory(df2)

        self.progress_updated.emit(30, "正在分块处理匹配源文件...")

        # 分块处理文件1
        chunks_processed = 0
        total_chunks = (total_rows1 - 1) // self.chunk_size + 1

        results = []

        # 获取文件1的headers
        headers1 = self._get_file_headers(self.file1_path, self.sheet1_name, file1_ext)

        # 使用多进程处理
        with ProcessPoolExecutor(max_workers=self.process_count) as executor:
            futures = []

            for start_row in range(1, total_rows1, self.chunk_size):
                future = executor.submit(
                    process_chunk_with_merge,
                    self.file1_path, self.sheet1_name, start_row, self.chunk_size,
                    file1_ext, headers1, df2, self.encoding, self.match_columns, self.output_columns
                )
                futures.append(future)

            # 收集结果
            for future in futures:
                try:
                    chunk_result = future.result(timeout=300)  # 5分钟超时
                    if not chunk_result.empty:
                        results.append(chunk_result)

                    chunks_processed += 1
                    progress = 30 + (chunks_processed / total_chunks) * 50
                    self.progress_updated.emit(int(progress),
                                            f"已处理 {chunks_processed}/{total_chunks} 个数据块")
                except Exception as e:
                    print(f"处理数据块时出错: {e}")

        self.progress_updated.emit(85, "正在合并处理结果...")

        # 合并所有结果
        if results:
            final_result = pd.concat(results, ignore_index=True)
            # 去除重复行
            final_result = final_result.drop_duplicates()
        else:
            final_result = pd.DataFrame()

        self.progress_updated.emit(95, "正在导出最终结果...")

        # 导出结果
        output_path = self._export_result(final_result)
        self.finished.emit(output_path)

    def _get_file_headers(self, file_path, sheet_name, file_ext):
        """获取文件的表头"""
        if file_ext == '.csv':
            # 对于CSV文件，读取第一行获取表头
            df = pd.read_csv(file_path, encoding=self.encoding, nrows=0)
            return df.columns.tolist()
        else:
            chunk = read_excel_chunk(file_path, sheet_name, 0, 1, file_ext, encoding=self.encoding)
            return chunk.columns.tolist() if not chunk.empty else []

    def _process_chunk_wrapper(self, file1_path, sheet1_name, start_row, chunk_size,
                              file1_ext, headers1, df2):
        """处理单个数据块的包装函数"""
        try:
            # 读取数据块
            chunk1 = read_excel_chunk(file1_path, sheet1_name, start_row, chunk_size,
                                    file1_ext, headers1, self.encoding)

            if chunk1.empty:
                return pd.DataFrame()

            # 优化内存使用
            chunk1 = optimize_dataframe_memory(chunk1)

            # 执行匹配
            result = self._perform_merge(chunk1, df2)

            # 清理内存
            del chunk1
            gc.collect()

            return result
        except Exception as e:
            print(f"处理数据块出错: {e}")
            return pd.DataFrame()

    def _read_complete_file(self, file_path, sheet_name, file_ext):
        """读取完整文件"""
        if file_ext == '.csv':
            # 读取CSV文件
            return pd.read_csv(file_path, encoding=self.encoding)
        elif file_ext == '.xls':
            # 使用xlrd处理xls文件
            encodings = ['gb18030', 'gbk', 'gb2312', 'utf-8']
            workbook = None

            for encoding in encodings:
                try:
                    workbook = xlrd.open_workbook(file_path, encoding_override=encoding)
                    break
                except Exception:
                    continue

            if not workbook:
                workbook = xlrd.open_workbook(file_path)

            sheet = workbook.sheet_by_name(sheet_name)

            # 创建DataFrame
            data = []
            headers = []

            # 获取表头
            for col in range(sheet.ncols):
                headers.append(str(sheet.cell_value(0, col)))

            # 获取数据
            for row in range(1, sheet.nrows):
                row_data = []
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    # 处理日期类型
                    if sheet.cell_type(row, col) == xlrd.XL_CELL_DATE:
                        try:
                            cell_value = xlrd.xldate.xldate_as_datetime(cell_value, workbook.datemode).strftime('%Y-%m-%d')
                        except:
                            pass
                    row_data.append(cell_value)
                data.append(row_data)

            return pd.DataFrame(data, columns=headers)
        else:
            return pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name)

    def _perform_merge(self, df1, df2):
        """执行数据合并"""
        # 创建匹配条件
        rename_dict = {}
        merge_on = []

        for col1, col2 in self.match_columns:
            col1_name = col1.split('-')[1]
            col2_name = col2.split('-')[1]
            rename_dict[col2_name] = col1_name
            merge_on.append(col1_name)

        # 只选择需要的列进行合并，减少内存使用和处理时间
        # 从df1中选择匹配列和用户选择的输出列
        df1_needed_columns = set(merge_on)
        for col in self.output_columns:
            if col in df1.columns:
                df1_needed_columns.add(col)

        # 从df2中选择匹配列和用户选择的输出列
        df2_needed_columns = set()
        for col1, col2 in self.match_columns:
            col2_name = col2.split('-')[1]
            df2_needed_columns.add(col2_name)

        for col in self.output_columns:
            if col in df2.columns:
                df2_needed_columns.add(col)

        # 只保留需要的列
        df1_filtered = df1[list(df1_needed_columns)].copy()
        df2_filtered = df2[list(df2_needed_columns)].copy()

            # 为df2的非匹配列添加后缀以避免同名冲突
        df2_rename_dict = {}
        for col in df2_filtered.columns:
            if col in rename_dict:
                df2_rename_dict[col] = rename_dict[col]
            elif col not in merge_on:
                df2_rename_dict[col] = f"{col}_被匹配文件"

        # 重命名df2的列
        df2_renamed = df2_filtered.rename(columns=df2_rename_dict)

        # 合并数据
        result = pd.merge(df1_filtered, df2_renamed, on=merge_on, how='inner')

        # 准备最终输出的列（只包含匹配列和用户选择的列）
        final_columns = merge_on.copy()

        # 添加用户选中的输出列
        for col in self.output_columns:
            if col in result.columns and col not in merge_on:
                final_columns.append(col)
            elif f"{col}_被匹配文件" in result.columns:
                final_columns.append(f"{col}_被匹配文件")

        # 选择最终输出的列
        return result[final_columns] if final_columns else result

    def _export_result(self, result):
        """导出结果到Excel文件（优化性能版本）"""
        # 获取输出文件路径
        output_dir = os.path.dirname(self.file2_path)
        base_output_path = os.path.join(output_dir, f"{self.output_filename}.xlsx")

        # 检查是否存在同名文件，如果存在则添加时间后缀
        if os.path.exists(base_output_path):
            current_time = datetime.datetime.now().strftime("%m%d_%H%M%S")
            output_path = os.path.join(output_dir, f"{self.output_filename}_{current_time}.xlsx")
        else:
            output_path = base_output_path

        # 确保路径使用正确的分隔符
        output_path = os.path.normpath(output_path)

        # 优化导出性能
        total_rows = len(result)

        if total_rows == 0:
            # 如果没有数据，创建空文件
            empty_df = pd.DataFrame(columns=result.columns if hasattr(result, 'columns') else [])
            empty_df.to_excel(output_path, index=False, engine='openpyxl')
            return output_path

        # 根据数据量选择导出策略
        if total_rows <= 100000:
            # 小数据量直接导出
            self.progress_updated.emit(96, f"正在导出 {total_rows:,} 行数据...")
            result.to_excel(output_path, index=False, engine='openpyxl')
            self.progress_updated.emit(100, "导出完成")
        else:
            # 大数据量分批导出
            self.progress_updated.emit(95, f"正在分批导出 {total_rows:,} 行数据...")

            # 使用openpyxl直接写入，性能更好
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "匹配结果"

            # 分批写入数据
            batch_size = 10000
            total_batches = (total_rows - 1) // batch_size + 1

            # 写入表头
            for r in dataframe_to_rows(result.head(0), index=False, header=True):
                ws.append(r)
                break  # 只写入表头

            # 分批写入数据
            for batch_num in range(total_batches):
                start_idx = batch_num * batch_size
                end_idx = min((batch_num + 1) * batch_size, total_rows)

                batch_data = result.iloc[start_idx:end_idx]

                # 写入当前批次数据
                for r in dataframe_to_rows(batch_data, index=False, header=False):
                    ws.append(r)

                # 更新进度
                progress = 95 + int((batch_num + 1) / total_batches * 4)
                self.progress_updated.emit(progress,
                                         f"正在导出第 {batch_num + 1}/{total_batches} 批数据 ({end_idx:,}/{total_rows:,} 行)")

            # 保存文件
            self.progress_updated.emit(99, "正在保存文件...")
            wb.save(output_path)
            wb.close()
            self.progress_updated.emit(100, "导出完成")

        return output_path


def read_excel_chunk(file_path, sheet_name, start_row, chunk_size, file_ext, headers=None, encoding='gbk'):
    """
    分块读取Excel/CSV文件的辅助函数
    """
    try:
        if file_ext == '.csv':
            # 处理CSV文件的分块读取
            if start_row == 0:
                # 读取表头和第一块数据
                return pd.read_csv(file_path, encoding=encoding, nrows=chunk_size)
            else:
                # 跳过表头，读取指定行数的数据
                # 注意：skiprows包含表头，所以需要跳过start_row行
                return pd.read_csv(file_path, encoding=encoding, skiprows=start_row, nrows=chunk_size, names=headers)
        elif file_ext == '.xls':
            # 使用xlrd处理xls文件
            encodings = ['gb18030', 'gbk', 'gb2312', 'utf-8']
            workbook = None

            for encoding in encodings:
                try:
                    workbook = xlrd.open_workbook(file_path, encoding_override=encoding)
                    break
                except Exception:
                    continue

            if not workbook:
                workbook = xlrd.open_workbook(file_path)

            sheet = workbook.sheet_by_name(sheet_name)

            # 如果没有提供headers，读取第一行作为headers
            if headers is None:
                headers = []
                for col in range(sheet.ncols):
                    headers.append(str(sheet.cell_value(0, col)))

            # 读取数据块
            data = []
            end_row = min(start_row + chunk_size, sheet.nrows)

            for row in range(start_row, end_row):
                row_data = []
                for col in range(sheet.ncols):
                    cell_value = sheet.cell_value(row, col)
                    # 处理日期类型
                    if sheet.cell_type(row, col) == xlrd.XL_CELL_DATE:
                        try:
                            cell_value = xlrd.xldate.xldate_as_datetime(cell_value, workbook.datemode).strftime('%Y-%m-%d')
                        except:
                            pass
                    row_data.append(cell_value)
                data.append(row_data)

            if data:
                return pd.DataFrame(data, columns=headers)
            else:
                return pd.DataFrame(columns=headers)
        else:
            # 使用pandas读取xlsx文件
            return pd.read_excel(file_path, engine='openpyxl', sheet_name=sheet_name,
                               skiprows=start_row-1 if start_row > 1 else 0, nrows=chunk_size)
    except Exception as e:
        print(f"读取文件块时出错: {e}")
        return pd.DataFrame()


def process_chunk_with_merge(file1_path, sheet1_name, start_row, chunk_size,
                           file1_ext, headers1, df2, encoding, match_columns, output_columns=None):
    """处理单个数据块并执行合并（支持编码）"""
    try:
        # 读取数据块
        chunk1 = read_excel_chunk(file1_path, sheet1_name, start_row, chunk_size,
                                file1_ext, headers1, encoding)

        if chunk1.empty:
            return pd.DataFrame()

        # 优化内存使用
        chunk1 = optimize_dataframe_memory(chunk1)

        # 执行合并逻辑
        # 创建匹配条件
        rename_dict = {}
        merge_on = []

        for col1, col2 in match_columns:
            col1_name = col1.split('-')[1]
            col2_name = col2.split('-')[1]
            rename_dict[col2_name] = col1_name
            merge_on.append(col1_name)

        # 只选择需要的列进行合并，减少内存使用和处理时间
        if output_columns is None:
            output_columns = []

        # 从chunk1中选择匹配列和用户选择的输出列
        chunk1_needed_columns = set(merge_on)
        for col in output_columns:
            if col in chunk1.columns:
                chunk1_needed_columns.add(col)

        # 从df2中选择匹配列和用户选择的输出列
        df2_needed_columns = set()
        for col1, col2 in match_columns:
            col2_name = col2.split('-')[1]
            df2_needed_columns.add(col2_name)

        for col in output_columns:
            if col in df2.columns:
                df2_needed_columns.add(col)

        # 只保留需要的列
        chunk1_filtered = chunk1[list(chunk1_needed_columns)].copy()
        df2_filtered = df2[list(df2_needed_columns)].copy()

        # 为df2的非匹配列添加后缀以避免同名冲突
        df2_rename_dict = {}
        for col in df2_filtered.columns:
            if col in rename_dict:
                df2_rename_dict[col] = rename_dict[col]
            elif col not in merge_on:
                df2_rename_dict[col] = f"{col}_被匹配文件"

        # 重命名df2的列
        df2_renamed = df2_filtered.rename(columns=df2_rename_dict)

        # 合并数据
        result = pd.merge(chunk1_filtered, df2_renamed, on=merge_on, how='inner')

        # 准备最终输出的列（只包含匹配列和用户选择的列）
        final_columns = merge_on.copy()
        for col in output_columns:
            if col in result.columns and col not in merge_on:
                final_columns.append(col)
            elif f"{col}_被匹配文件" in result.columns:
                final_columns.append(f"{col}_被匹配文件")

        # 选择最终输出的列
        result = result[final_columns] if final_columns else result

        # 清理内存
        del chunk1, chunk1_filtered
        gc.collect()

        return result
    except Exception as e:
        print(f"处理数据块出错: {e}")
        return pd.DataFrame()


def process_chunk_merge(chunk1, chunk2, merge_on, rename_dict):
    """
    处理单个数据块的合并
    """
    try:
        # 重命名df2的列
        chunk2_renamed = chunk2.rename(columns=rename_dict)

        # 合并数据
        result = pd.merge(chunk1, chunk2_renamed, on=merge_on, how='inner')
        return result
    except Exception as e:
        print(f"合并数据块时出错: {e}")
        return pd.DataFrame()


def optimize_dataframe_memory(df):
    """
    优化DataFrame的内存使用
    """
    for col in df.columns:
        col_type = df[col].dtype

        if col_type != 'object':
            c_min = df[col].min()
            c_max = df[col].max()

            if str(col_type)[:3] == 'int':
                if c_min > np.iinfo(np.int8).min and c_max < np.iinfo(np.int8).max:
                    df[col] = df[col].astype(np.int8)
                elif c_min > np.iinfo(np.int16).min and c_max < np.iinfo(np.int16).max:
                    df[col] = df[col].astype(np.int16)
                elif c_min > np.iinfo(np.int32).min and c_max < np.iinfo(np.int32).max:
                    df[col] = df[col].astype(np.int32)

            elif str(col_type)[:5] == 'float':
                if c_min > np.finfo(np.float32).min and c_max < np.finfo(np.float32).max:
                    df[col] = df[col].astype(np.float32)

    return df


def main():
    app = QApplication(sys.argv)
    window = ExcelCompareTool()
    window.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
